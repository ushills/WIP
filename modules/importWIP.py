# import WIP data from WIP file

# import libraries
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from colorama import init, Fore

# initialise colorama
init()
import sqlite3
import os
import re

# supress warnings
import warnings

warnings.filterwarnings("ignore")


# import_wip_data takes information in the form
# import_wip_data(database_name, importdirectory)
def import_wip_data(database_name, directory_name):
    print("importing from", directory_name, "to", database_name)

    # first extract the files from the _directory
    file_list = list_files(os.path.normpath(directory_name))

    # check the files are wip files
    wip_files = check_wipfile(file_list)

    # extract the data from the excel worksheet cells
    # and export it to the sqldatabase
    for wip_filename in wip_files:
        wip_data = import_data(wip_filename)
        export_data_sql(wip_data, database_name)


# search through _directory to only return excel files
def list_files(directory):
    print("searching for files.....")
    file_list = []
    for root, directories, filenames in os.walk(os.path.normpath(directory)):
        for filename in filenames:
            raw_filename = str(os.path.join(root, filename))
            # skip temp files beginning with ~
            if re.search(r"(~.*).*", raw_filename):
                continue
            # print raw_filename
            # check if the file is excel, i.e. ends .xlsx
            if re.search(r"(.*).xls?", raw_filename):
                # print 'Appending', raw_filename
                file_list.append(raw_filename)
    print("found", len(file_list), "files")
    return file_list


# now check that the excel file is a wip file
def check_wipfile(filelist):
    filtered_filelist = []
    print("filtering files.....")
    worksheet_name = "Executive Summary"
    project_name_cell = "A5"
    project_number_cell = "B6"

    # go through the filelist and check each file
    for excel_file in filelist:
        # Try to open the WIP excel file
        try:
            wip_workbook = load_workbook(
                filename=excel_file, read_only=True, data_only=True
            )
            # and try to open the worksheet
            wip_worksheet = wip_workbook[worksheet_name]
            # print 'checking file', excel_file
        except InvalidFileException:
            continue
        except BaseException:
            continue
        if wip_worksheet[project_name_cell].value != "Project Name:":
            # print excel_file, 'is not a wip file'
            continue
        elif wip_worksheet[project_number_cell].value is None:
            continue
        else:
            filtered_filelist.append(excel_file)
            # print 'adding', excel_file, 'to wipfile list'
    print("filtered", len(filelist), "files to", len(filtered_filelist), "files")
    return filtered_filelist


# function to import wip data from excel file
def import_data(wip_filename):

    # set variables for location of various fields
    # e.g. forecast sale, forecast cost, contribution, cost to date etc
    # excel cell references in the following format
    # variable = (excel row, excel column)
    # we will return the data as a dictionary
    # with the variable wip_data
    wip_data = dict()

    # header variables
    worksheet_name = "Executive Summary"
    project_name_ref = "B5"
    project_number_ref = "B6"
    wip_date_ref = "B7"

    # columns for this month and Now
    now_column = "C"
    this_month_column = "F"

    # sales account variables
    agreed_variations_no_ref = now_column + "20"
    budget_variations_no_ref = now_column + "21"
    submitted_variations_no_ref = now_column + "22"
    order_value_ref = this_month_column + "18"
    agreed_variations_value_ref = this_month_column + "20"
    budget_variations_value_ref = this_month_column + "21"
    submitted_variations_value_ref = this_month_column + "22"

    # Reserve variable
    reserve_agreed_ref = this_month_column + "30"
    reserve_budget_ref = this_month_column + "31"
    reserve_submitted_ref = this_month_column + "32"
    contracharges_ref = this_month_column + "36"

    # cost variables
    current_cost_ref = this_month_column + "43"
    cost_to_complete_ref = this_month_column + "44"
    defect_provision_ref = this_month_column + "48"

    # margin variables
    contract_contribution_ref = this_month_column + "55"
    betterments_risks_ref = this_month_column + "56"
    managers_view_ref = this_month_column + "57"

    # total variables
    sale_subtotal_ref = this_month_column + "26"
    # reserve_subtotal_ref = this_month_column + "34"
    forecast_sale_total_ref = this_month_column + "38"
    variations_no_total_ref = now_column + "24"
    forecast_cost_total_ref = this_month_column + "50"
    forecast_margin_total_ref = this_month_column + "59"

    # cash recovery variables
    latest_application_ref = this_month_column + "67"
    total_certified_ref = this_month_column + "72"

    # Extract the data from the WIP excel spreadsheet and set variables

    # Try to open the WIP excel file
    try:
        wip_workbook = load_workbook(
            filename=wip_filename, read_only=True, data_only=True
        )
        # and try to open the worksheet
        print("Opening workbook", worksheet_name, "in", wip_filename)
    except FileNotFoundError:
        print(Fore.RED)
        print(wip_filename, " does not exist, exiting")
        print(Fore.RESET)
        quit()

    # extract the cell information
    # using the format wip_data['field'] = wip_worksheet([fieldRef].value)
    wip_worksheet = wip_workbook[worksheet_name]
    wip_data["projectName"] = wip_worksheet[project_name_ref].value
    wip_data["projectNumber"] = wip_worksheet[project_number_ref].value
    wip_data["wipDate"] = str(wip_worksheet[wip_date_ref].value)
    wip_data["wipDate"] = (wip_data["wipDate"].rsplit(" "))[0]
    wip_date_formatted = wip_data["wipDate"]
    print(
        "Extracting",
        wip_data["projectNumber"],
        "-",
        wip_data["projectName"],
        "data for",
        wip_date_formatted,
        "\n",
    )

    # check that the date is not None
    try:
        assert wip_data["wipDate"] != "None", (
            wip_filename
            + " contains an incorrect date format of None\nPlease correct and retry"
        )
        print("Date...........okay")
    except AssertionError as e:
        print(Fore.RED)
        print("-" * 80)
        print("FATAL...\n", e)
        print("-" * 80)
        print(Fore.RESET)
        quit()

    # extract variation information
    wip_data["agreedVariationsNo"] = int(wip_worksheet[agreed_variations_no_ref].value)
    wip_data["budgetVariationsNo"] = int(wip_worksheet[budget_variations_no_ref].value)
    wip_data["submittedVariationsNo"] = int(
        wip_worksheet[submitted_variations_no_ref].value
    )
    wip_data["variationsNoTotal"] = int(wip_worksheet[variations_no_total_ref].value)
    variations_no_total_check = (
        wip_data["agreedVariationsNo"]
        + wip_data["budgetVariationsNo"]
        + wip_data["submittedVariationsNo"]
    )
    # check variations no balance
    if wip_data["variationsNoTotal"] != variations_no_total_check:
        print(Fore.RED)
        print("Number of variations incorrect")
        print(Fore.RESET)
        quit()
    else:
        print("Variations.....okay")

    # extract sales information
    wip_data["orderValue"] = int(wip_worksheet[order_value_ref].value)
    wip_data["agreedVariationsValue"] = int(
        wip_worksheet[agreed_variations_value_ref].value
    )
    wip_data["budgetVariationsValue"] = int(
        wip_worksheet[budget_variations_value_ref].value
    )
    wip_data["submittedVariationsValue"] = int(
        wip_worksheet[submitted_variations_value_ref].value
    )
    wip_data["saleSubtotal"] = int(wip_worksheet[sale_subtotal_ref].value)
    sale_subtotal_check = (
        wip_data["orderValue"]
        + wip_data["agreedVariationsValue"]
        + wip_data["budgetVariationsValue"]
        + wip_data["submittedVariationsValue"]
    )
    # check sales values balance
    if abs(sale_subtotal_check - wip_data["saleSubtotal"]) > 5:
        print(Fore.RED)
        print("Sales value incorrect")
        print(Fore.RESET)
        quit()
    else:
        print("Sales.....okay")

    # extract reserve information
    wip_data["reserveAgreed"] = int(wip_worksheet[reserve_agreed_ref].value)
    wip_data["reserveBudget"] = int(wip_worksheet[reserve_budget_ref].value)
    wip_data["reserveSubmitted"] = int(wip_worksheet[reserve_submitted_ref].value)
    wip_data["contracharges"] = int(wip_worksheet[contracharges_ref].value)
    wip_data["forecastSaleTotal"] = int(wip_worksheet[forecast_sale_total_ref].value)
    sum_checksale = (
        wip_data["orderValue"]
        + wip_data["agreedVariationsValue"]
        + wip_data["budgetVariationsValue"]
        + wip_data["submittedVariationsValue"]
        + wip_data["reserveAgreed"]
        + wip_data["reserveBudget"]
        + wip_data["reserveSubmitted"]
        + wip_data["contracharges"]
    )
    if abs(sum_checksale - wip_data["forecastSaleTotal"]) > 5:
        print(Fore.RED)
        print("Forecast sale subtotal incorrect")
        print(Fore.RESET)
        quit()
    else:
        print("Forecast sale.....okay")

    # extract cost information
    wip_data["currentCost"] = int(wip_worksheet[current_cost_ref].value)
    wip_data["costToComplete"] = int(wip_worksheet[cost_to_complete_ref].value)
    try:
        wip_data["defectProvision"] = int(wip_worksheet[defect_provision_ref].value)
    except TypeError:
        wip_data["defectProvision"] = 0
    wip_data["forecastCostTotal"] = int(wip_worksheet[forecast_cost_total_ref].value)
    sum_check_cost = (
        wip_data["currentCost"]
        + wip_data["costToComplete"]
        + wip_data["defectProvision"]
    )
    if abs(sum_check_cost - wip_data["forecastCostTotal"]) > 5:
        print(Fore.RED)
        print("Forecast cost subtotal incorrect")
        print(Fore.RESET)
        quit()
    else:
        print("Forecast cost.....okay")

    # extract margin information
    wip_data["contractContribution"] = int(
        wip_worksheet[contract_contribution_ref].value
    )
    wip_data["bettermentsRisks"] = int(wip_worksheet[betterments_risks_ref].value)

    # check if the managersView is blank and would cause an error
    # if true add a 0
    try:
        wip_data["managersView"] = int(wip_worksheet[managers_view_ref].value)
    except TypeError:
        wip_data["managersView"] = 0
    # except ValueError:
    #     wip_data["managersView"] = 0

    wip_data["forecastMarginTotal"] = int(
        wip_worksheet[forecast_margin_total_ref].value
    )
    sum_check_margin = (
        wip_data["contractContribution"]
        + wip_data["bettermentsRisks"]
        + wip_data["managersView"]
    )
    if abs(sum_check_margin - wip_data["forecastMarginTotal"]) > 5:
        print(Fore.RED)
        print("Forecast margin subtotal incorrect")
        print(Fore.RESET)
        quit()
    else:
        print("Forecast margin.....okay")

    # extract cash information
    # if the field is blank insert a 0
    try:
        wip_data["latestApplication"] = int(wip_worksheet[latest_application_ref].value)
    except TypeError:
        wip_data["latestApplication"] = 0
    try:
        wip_data["totalCertified"] = int(wip_worksheet[total_certified_ref].value)
    except TypeError:
        wip_data["totalCertified"] = 0

    print(Fore.GREEN)
    print("All data imported sucessfully\n")
    print(Fore.RESET)
    return wip_data


# function to import data into an SQL database
def export_data_sql(data_list, database):
    # print data_list
    project_name = data_list["projectName"]
    # wip_date = data_list["wipDate"]
    conn = sqlite3.connect(os.path.normpath(database))
    cur = conn.cursor()

    # delete the database table if it exists...for testing only
    # cur.execute('''DROP TABLE IF EXISTS wipdata''')

    # create job name table
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS projectname (
        id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
        name TEXT UNIQUE
        )"""
    )

    # create wipdata table if doesn't exist
    cur.execute(
        """
    CREATE TABLE IF NOT EXISTS wipdata (
        projectNumber TEXT, projectName INTEGER, wipDate TEXT,
        agreedVariationsNo INTEGER, budgetVariationsNo INTEGER,
        submittedVariationsNo INTEGER, variationsNoTotal INTEGER,
        orderValue INTEGER, agreedVariationsValue INTEGER,
        budgetVariationsValue INTEGER, submittedVariationsValue INTEGER,
        saleSubtotal INTEGER, reserveAgreed INTEGER, reserveBudget INTEGER,
        reserveSubmitted INTEGER, contracharges INTEGER,
        forecastSaleTotal INTEGER, currentCost INTEGER, costToComplete INTEGER,
        defectProvision INTEGER, forecastCostTotal INTEGER,
        contractContribution INTEGER, bettermentsRisks INTEGER,
        managersView INTEGER, forecastMarginTotal INTEGER,
        latestApplication INTEGER, totalCertified INTEGER,
        PRIMARY KEY (projectNumber, wipDate));"""
    )

    # check if the project name exists in the table project name
    cur.execute(
        """INSERT OR IGNORE INTO projectname (name)
        VALUES ( ? )""",
        (project_name,),
    )
    cur.execute("SELECT id FROM projectname WHERE name = ?", (project_name,))
    project_name_id = cur.fetchone()[0]
    data_list["projectName"] = project_name_id

    # run through the data and allocate to each field
    cur.executemany(
        """INSERT OR REPLACE INTO wipdata
        (projectNumber, projectName, wipDate, agreedVariationsNo,
        budgetVariationsNo, submittedVariationsNo, variationsNoTotal,
        orderValue, agreedVariationsValue, budgetVariationsValue,
        submittedVariationsValue, saleSubtotal, reserveAgreed,
        reserveBudget, reserveSubmitted, contracharges,
        forecastSaleTotal, currentCost, costToComplete,
        defectProvision, forecastCostTotal, contractContribution,
        bettermentsRisks, managersView, forecastMarginTotal,
        latestApplication, totalCertified)
    VALUES
        (:projectNumber, :projectName, :wipDate, :agreedVariationsNo,
        :budgetVariationsNo, :submittedVariationsNo, :variationsNoTotal,
        :orderValue, :agreedVariationsValue, :budgetVariationsValue,
        :submittedVariationsValue, :saleSubtotal, :reserveAgreed,
        :reserveBudget, :reserveSubmitted, :contracharges,
        :forecastSaleTotal, :currentCost, :costToComplete,
        :defectProvision, :forecastCostTotal, :contractContribution,
        :bettermentsRisks, :managersView, :forecastMarginTotal,
        :latestApplication, :totalCertified)""",
        [data_list],
    )

    # Commit the changes to the database
    conn.commit()
    cur.close()
    conn.close()
